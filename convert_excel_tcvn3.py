# -*- coding: utf-8 -*-
from __future__ import annotations
import json
import re
import subprocess
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Tuple
from dataclasses import dataclass
from datetime import datetime

import pandas as pd

# Đường dẫn đến file map
MAP_JSON = Path(__file__).parent / "tcvn3_map.json"
MAP_CSV = Path(__file__).parent / "tcvn3_map.csv"
BUILD_SCRIPT = Path(__file__).parent / "build_tcvn3_map.py"

# Biến global để cache map
_TCVN3_TO_UNI: Dict[str, str] = {}
_TCVN3_REGEX = None

# Tập ký tự tiếng Việt hợp lệ (Latin + dấu chuẩn + số, khoảng trắng, punctuation phổ biến)
_VIET_UNI_OK = set(
    "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZàáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡ"
    "ùúụủũưừứựửữỳýỵỷỹđÀÁẠẢÃÂẦẤẬẨẪĂẰẮẶẲẴÈÉẸẺẼÊỀẾỆỂỄÌÍỊỈĨ"
    "ÒÓỌỎÕÔỒỐỘỔỖƠỜỚỢỞỠÙÚỤỦŨƯỪỨỰỬỮỲÝỴỶỸĐ"
    " 0123456789.,;:!?\"'()[]{}-_/\\|@#%&*+=…""''\n\t€$¥£₫%‰°±×÷"
)


@dataclass
class ConversionLog:
    """Log thông tin một cell được convert"""
    sheet: str
    row: int
    col: int
    col_name: str
    original: str
    converted: str
    was_unicode: bool


@dataclass
class ConversionStats:
    """Thống kê quá trình convert"""
    total_cells: int = 0
    string_cells: int = 0
    already_unicode: int = 0
    converted_cells: int = 0
    unchanged_cells: int = 0
    sheets_processed: int = 0
    logs: List[ConversionLog] = None
    
    def __post_init__(self):
        if self.logs is None:
            self.logs = []


def load_tcvn3_map() -> Dict[str, str]:
    """
    Load bảng map TCVN3 -> Unicode từ file JSON hoặc CSV.
    Nếu file không tồn tại, tự động chạy build_tcvn3_map.py để tạo.
    """
    global _TCVN3_TO_UNI, _TCVN3_REGEX
    
    # Nếu đã load rồi, return luôn
    if _TCVN3_TO_UNI:
        return _TCVN3_TO_UNI
    
    # Ưu tiên load từ JSON
    if MAP_JSON.exists():
        try:
            with MAP_JSON.open("r", encoding="utf-8") as f:
                _TCVN3_TO_UNI = json.load(f)
                print(f"✅ Đã tải {len(_TCVN3_TO_UNI)} mapping từ {MAP_JSON.name}")
        except Exception as e:
            print(f"⚠️ Không đọc được {MAP_JSON.name}: {e}")
            _TCVN3_TO_UNI = {}
    # Nếu không có JSON, thử load từ CSV
    elif MAP_CSV.exists():
        try:
            df = pd.read_csv(MAP_CSV, encoding="utf-8")
            _TCVN3_TO_UNI = dict(zip(df["TCVN3"], df["UNICODE"]))
            print(f"✅ Đã tải {len(_TCVN3_TO_UNI)} mapping từ {MAP_CSV.name}")
        except Exception as e:
            print(f"⚠️ Không đọc được {MAP_CSV.name}: {e}")
            _TCVN3_TO_UNI = {}
    
    # Nếu vẫn không có map, tự động build
    if not _TCVN3_TO_UNI:
        print("📦 Không tìm thấy file map. Đang tạo map từ trang web...")
        if BUILD_SCRIPT.exists():
            try:
                # Chạy build_tcvn3_map.py --build
                result = subprocess.run(
                    [sys.executable, str(BUILD_SCRIPT), "--build"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                )
                if result.returncode == 0:
                    print(result.stdout)
                    # Thử load lại
                    if MAP_JSON.exists():
                        with MAP_JSON.open("r", encoding="utf-8") as f:
                            _TCVN3_TO_UNI = json.load(f)
                            print(f"✅ Đã tải {len(_TCVN3_TO_UNI)} mapping từ {MAP_JSON.name}")
                    elif MAP_CSV.exists():
                        df = pd.read_csv(MAP_CSV, encoding="utf-8")
                        _TCVN3_TO_UNI = dict(zip(df["TCVN3"], df["UNICODE"]))
                        print(f"✅ Đã tải {len(_TCVN3_TO_UNI)} mapping từ {MAP_CSV.name}")
                else:
                    print(f"❌ Lỗi khi build map: {result.stderr}")
                    raise RuntimeError("Không thể tạo map TCVN3. Hãy chạy build_tcvn3_map.py --build thủ công.")
            except Exception as e:
                print(f"❌ Lỗi khi chạy build script: {e}")
                raise RuntimeError("Không thể tạo map TCVN3. Hãy chạy build_tcvn3_map.py --build thủ công.")
        else:
            raise FileNotFoundError(
                f"Không tìm thấy file map ({MAP_JSON.name} hoặc {MAP_CSV.name}) "
                f"và không tìm thấy script build ({BUILD_SCRIPT.name})"
            )
    
    # Tạo regex pattern từ keys
    if _TCVN3_TO_UNI:
        _TCVN3_REGEX = re.compile("|".join(map(re.escape, _TCVN3_TO_UNI.keys())))
    
    return _TCVN3_TO_UNI


def looks_like_unicode_vietnamese(s: str) -> bool:
    """
    Kiểm tra xem chuỗi có phải là tiếng Việt Unicode hợp lệ hay không.
    
    LOGIC TỐI ƯU v2.1:
    1. Empty/whitespace → TRUE (bỏ qua)
    2. Chỉ số + dấu (VD: "123", "---", "2024-11-09") → TRUE (không cần review)
    3. Có chữ cái VN Unicode + không có ký tự lạ → TRUE
    4. Có ký tự lạ (TCVN3) → FALSE (cần convert)
    
    Returns:
        True nếu chuỗi đã là Unicode Việt hợp lệ (bỏ qua không cần convert)
        False nếu có ký tự lạ (có thể là TCVN3)
    """
    if not s:
        return True
    
    s_stripped = s.strip()
    if not s_stripped:
        return True  # Chỉ whitespace
    
    # Quick check: Chỉ có số, dấu câu cơ bản (không có chữ)
    # VD: "123", "---", "...", "2024-11-09", "1,234.56"
    has_letter = False
    
    for ch in s:
        if ch.isalpha():
            has_letter = True
            break
    
    # Nếu không có chữ cái → OK (số, dấu, date...)
    if not has_letter:
        return True
    
    # Có chữ cái → Check kỹ hơn
    for ch in s:
        # Cho qua nếu trong whitelist
        if ch in _VIET_UNI_OK:
            continue
        # Hoặc là ký tự có category dấu/khoảng trắng/punct/symbol bình thường
        cat = unicodedata.category(ch)
        if cat.startswith(('Z', 'P', 'C', 'S')):  # Separator, Punctuation, Control, Symbol
            continue
        # Gặp ký tự lạ ngoài whitelist
        return False
    return True


def is_likely_non_text_content(s: str) -> bool:
    """
    Kiểm tra xem cell có phải là nội dung không phải text tiếng Việt.
    Dùng để filter ra các cell không cần review (số, date, dấu...)
    
    Returns:
        True nếu không cần review (số thuần, date, dấu câu...)
        False nếu cần review (có text chữ cái)
    """
    if not s or not s.strip():
        return True
    
    s_stripped = s.strip()
    
    # Chỉ có số + dấu phân cách
    if s_stripped.replace('.', '').replace(',', '').replace('-', '').replace('/', '').replace(':', '').isdigit():
        return True  # VD: "123", "2024-11-09", "1,234.56", "10:30"
    
    # Chỉ có dấu câu/ký hiệu (không có chữ, số)
    has_alnum = any(ch.isalnum() for ch in s_stripped)
    if not has_alnum:
        return True  # VD: "---", "...", "***", "- - -"
    
    # Có chữ cái hoặc chữ số mixed → cần review
    return False


def tcvn3_to_unicode(s: str) -> str:
    """
    Chuyển chuỗi từ mã TCVN3 (.VnTime) sang Unicode.
    Tự động load map nếu chưa load.
    """
    if not s:
        return s
    
    # Đảm bảo map đã được load
    if not _TCVN3_TO_UNI:
        load_tcvn3_map()
    
    if not _TCVN3_TO_UNI or not _TCVN3_REGEX:
        # Nếu vẫn không có map, return nguyên bản
        return s
    
    return _TCVN3_REGEX.sub(lambda m: _TCVN3_TO_UNI.get(m.group(0), m.group(0)), s)


def convert_excel(
    input_path: str | Path,
    output_path: str | Path,
    skip_unicode: bool = True,
    progress_callback=None,
    skip_selection: dict = None,
    highlight_converted: bool = False,
    highlight_color: str = "#FFFF00",
) -> ConversionStats:
    """
    Chuyển đổi file Excel từ TCVN3 sang Unicode với các tính năng nâng cao.
    
    Args:
        input_path: Đường dẫn đến file Excel input (TCVN3)
        output_path: Đường dẫn đến file Excel output (Unicode)
        skip_unicode: Nếu True, bỏ qua các cell đã là Unicode chuẩn
        progress_callback: Hàm callback(sheet_name, sheet_index, total_sheets) để báo tiến trình
        skip_selection: Dict[cell_id, should_skip] - Custom skip selection
        highlight_converted: Nếu True, đánh dấu màu cells đã convert
        highlight_color: Màu highlight (hex color)
        
    Returns:
        ConversionStats: Thống kê chi tiết quá trình convert
    """
    input_path = Path(input_path)
    output_path = Path(output_path)
    
    # Đảm bảo map đã được load (sẽ tự động build nếu chưa có)
    load_tcvn3_map()
    
    stats = ConversionStats()

    # Đọc toàn bộ sheets
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    
    xls = pd.ExcelFile(input_path, engine="openpyxl")
    out_writer = pd.ExcelWriter(output_path, engine="openpyxl")
    
    # For highlighting
    converted_cells_coords = []  # List of (sheet_name, row, col)

    total_sheets = len(xls.sheet_names)
    skip_selection = skip_selection or {}
    
    for sheet_idx, sheet in enumerate(xls.sheet_names):
        if progress_callback:
            progress_callback(sheet, sheet_idx, total_sheets)
            
        df = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=object)  # Không dùng header tự động
        stats.sheets_processed += 1

        # Xử lý từng cell (bao gồm cả dòng đầu tiên)
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                cell_value = df.iloc[row_idx, col_idx]
                stats.total_cells += 1
                
                # Chỉ xử lý các ô là chuỗi
                if isinstance(cell_value, str) and cell_value.strip():
                    stats.string_cells += 1
                    original = cell_value
                    
                    # Kiểm tra xem đã là Unicode chuẩn chưa
                    # Excel row is 1-indexed, so row_idx + 1
                    cell_id = f"{sheet}_{row_idx + 1}_{col_idx}"
                    is_unicode = looks_like_unicode_vietnamese(original)
                    
                    if skip_unicode and is_unicode:
                        # Check custom skip selection
                        should_skip = skip_selection.get(cell_id, True)
                        if should_skip:
                            stats.already_unicode += 1
                            continue
                        # Else: User wants to convert even Unicode cells
                    
                    # Convert
                    converted = tcvn3_to_unicode(original)
                    
                    # Log nếu có thay đổi
                    if converted != original:
                        stats.converted_cells += 1
                        df.iloc[row_idx, col_idx] = converted
                        
                        # Track for highlighting
                        if highlight_converted:
                            # Excel uses 1-indexed
                            converted_cells_coords.append((sheet, row_idx + 1, col_idx + 1))
                        
                        # Log chi tiết
                        log = ConversionLog(
                            sheet=sheet,
                            row=row_idx + 1,  # 1-indexed for Excel
                            col=col_idx,
                            col_name=f"Col_{col_idx}",  # Generic column name
                            original=original,
                            converted=converted,
                            was_unicode=is_unicode,
                        )
                        stats.logs.append(log)
                    else:
                        stats.unchanged_cells += 1

        df.to_excel(out_writer, sheet_name=sheet, index=False, header=False)

    out_writer.close()
    
    # Apply highlighting if requested
    if highlight_converted and converted_cells_coords:
        try:
            wb = load_workbook(output_path)
            fill = PatternFill(start_color=highlight_color.replace("#", ""),
                             end_color=highlight_color.replace("#", ""),
                             fill_type="solid")
            
            for sheet_name, row, col in converted_cells_coords:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
            
            wb.save(output_path)
            print(f"🎨 Đã đánh dấu {len(converted_cells_coords)} cells")
        except Exception as e:
            print(f"⚠️ Không thể đánh dấu màu: {e}")
    
    print(f"✅ Ghi xong: {output_path}")
    return stats


def preview_conversion(
    input_path: str | Path,
    max_samples: int = 9999999,
) -> List[ConversionLog]:
    """
    Xem trước các cell sẽ được convert mà không thực sự ghi file.
    
    Args:
        input_path: Đường dẫn đến file Excel input
        max_samples: Số lượng mẫu tối đa để hiển thị (None = tất cả)
        
    Returns:
        List[ConversionLog]: Danh sách các cell sẽ được convert
    """
    input_path = Path(input_path)
    load_tcvn3_map()
    
    samples = []
    xls = pd.ExcelFile(input_path, engine="openpyxl")
    
    for sheet in xls.sheet_names:
        # Không dùng header tự động để đọc cả dòng 1
        df = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=object)
        
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                # Kiểm tra giới hạn nếu có
                if max_samples is not None and len(samples) >= max_samples:
                    break
                    
                cell_value = df.iloc[row_idx, col_idx]
                
                if isinstance(cell_value, str) and cell_value.strip():
                    original = cell_value
                    is_unicode = looks_like_unicode_vietnamese(original)
                    converted = tcvn3_to_unicode(original)
                    
                    # Log TẤT CẢ các cell có text (bao gồm cả Unicode)
                    # để user có thể review đầy đủ
                    log = ConversionLog(
                        sheet=sheet,
                        row=row_idx + 1,  # 1-indexed for Excel
                        col=col_idx,
                        col_name=f"Col_{col_idx}",
                        original=original,
                        converted=converted,
                        was_unicode=is_unicode,
                    )
                    samples.append(log)
            
            # Break outer loop if we have enough samples
            if max_samples is not None and len(samples) >= max_samples:
                break
        
        # Break sheet loop if we have enough samples
        if max_samples is not None and len(samples) >= max_samples:
            break
    
    return samples


def export_conversion_log(stats: ConversionStats, log_path: str | Path) -> None:
    """
    Xuất log chi tiết ra file text.
    
    Args:
        stats: Thống kê conversion
        log_path: Đường dẫn file log output
    """
    log_path = Path(log_path)
    
    with log_path.open("w", encoding="utf-8") as f:
        f.write("=" * 80 + "\n")
        f.write(f"TCVN3 → Unicode Conversion Log\n")
        f.write(f"Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"📊 Thống kê:\n")
        f.write(f"  - Tổng số cells: {stats.total_cells:,}\n")
        f.write(f"  - Cells chứa text: {stats.string_cells:,}\n")
        f.write(f"  - Đã là Unicode chuẩn: {stats.already_unicode:,}\n")
        f.write(f"  - Đã convert: {stats.converted_cells:,}\n")
        f.write(f"  - Không đổi: {stats.unchanged_cells:,}\n")
        f.write(f"  - Số sheets: {stats.sheets_processed}\n\n")
        
        if stats.logs:
            f.write(f"📝 Chi tiết {len(stats.logs)} cells đã convert:\n")
            f.write("-" * 80 + "\n")
            
            for i, log in enumerate(stats.logs, 1):
                f.write(f"\n[{i}] Sheet: {log.sheet} | Row: {log.row} | Col: {log.col_name}\n")
                f.write(f"    BEFORE: {log.original}\n")
                f.write(f"    AFTER:  {log.converted}\n")
        else:
            f.write("Không có cell nào cần convert.\n")
    
    print(f"✅ Đã xuất log: {log_path}")


if __name__ == "__main__":
    # Ví dụ đường dẫn, đổi lại cho phù hợp
    convert_excel(
        r"D:\K\Code\Python\excel_tcvn3_converter\input_tcvn3.xlsx",
        r"D:\K\Code\Python\excel_tcvn3_converter\output_unicode.xlsx",
    )
